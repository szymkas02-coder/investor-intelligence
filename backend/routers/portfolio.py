"""
backend/routers/portfolio.py — Portfolio positions and transactions

GET  /portfolio              — current positions with live P&L
POST /portfolio/transaction  — record a buy/sell/dividend
GET  /portfolio/ike          — IKE contribution status for current year
POST /portfolio/upload-broker — AI-parsed broker Excel/CSV import
"""

import io
import json
import os
import re
import uuid
from datetime import date
from typing import Annotated

import openpyxl
import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from backend.auth import get_current_user, require_non_guest
from backend.database import get_db, get_db_write
from backend.models import (
    PortfolioResponse, Position, AccountSummary,
    TransactionCreate, TransactionResponse,
    TransactionRow, TransactionsResponse,
)

router = APIRouter(prefix="/portfolio", tags=["portfolio"])

# IKE annual contribution limits (PLN), set by Polish Ministry of Finance
IKE_LIMITS = {
    2016: 12165.0, 2017: 12789.0, 2018: 13329.0, 2019: 14295.0,
    2020: 15681.0, 2021: 15777.0, 2022: 17766.0,
    2023: 20805.0, 2024: 23472.0, 2025: 26019.0, 2026: 28260.0,
}

# IKZE annual contribution limits (PLN). IKZE = 1.2 × average monthly wage
# (1.8× for self-employed, not modelled — single-rate approximation).
IKZE_LIMITS = {
    2016: 4866.00, 2017: 5115.60, 2018: 5331.60, 2019: 5718.00,
    2020: 6272.40, 2021: 6310.80, 2022: 7106.40,
    2023: 8322.00, 2024: 9388.80, 2025: 10407.60, 2026: 11304.00,
}


def _compute_account_summary(db, user_id: str, year: int) -> list:
    """
    Aggregate this year's PLN contributions (BUY + DEPOSIT) per account_type
    directly from user_transactions. Returns list[AccountSummary].

    Why aggregate live instead of using the ike_contributions table?
    - The ike_contributions table only tracks IKE; mirroring it for IKZE and
      regular would need a parallel table and migration.
    - Live aggregation is naturally edit-safe: if a buy is modified or
      deleted the running totals stay correct without extra bookkeeping.
    - At single-user scale this is cheap (one indexed query per page load).
    """
    rows = db.execute("""
        SELECT account_type,
               COALESCE(SUM(CASE WHEN type = 'deposit' THEN price_pln
                                 WHEN type = 'buy'     THEN shares * price_pln
                                 ELSE 0 END), 0) AS contributed
        FROM user_transactions
        WHERE user_id = %s
          AND EXTRACT(YEAR FROM date) = %s
          AND type IN ('buy', 'deposit')
        GROUP BY account_type
    """, [user_id, year]).fetchall()

    by_type = {r[0]: float(r[1] or 0) for r in rows}

    summaries = []
    for account_type, limits_map in (("IKE", IKE_LIMITS), ("IKZE", IKZE_LIMITS), ("regular", None)):
        contributed = by_type.get(account_type, 0.0)
        if limits_map is not None:
            limit = limits_map.get(year)
            remaining = (limit - contributed) if limit is not None else None
        else:
            limit = None
            remaining = None
        summaries.append(AccountSummary(
            account_type = account_type,
            year         = year,
            contributed  = round(contributed, 2),
            limit        = limit,
            remaining    = round(remaining, 2) if remaining is not None else None,
        ))
    return summaries



@router.get("", response_model=PortfolioResponse)
def get_portfolio(
    db:      Annotated[object, Depends(get_db)],
    user_id: Annotated[str, Depends(get_current_user)],
):
    # Latest EURPLN and USDPLN for currency conversion
    fx_row = db.execute("""
        SELECT eurpln, usdpln FROM daily_features
        WHERE eurpln IS NOT NULL AND usdpln IS NOT NULL
        ORDER BY date DESC LIMIT 1
    """).fetchone()
    eurpln = float(fx_row[0]) if fx_row else 4.25
    usdpln = float(fx_row[1]) if fx_row else 3.85

    # GBP/PLN ≈ GBPUSD * USDPLN; approximate via EURPLN * 1.17 (rough)
    gbppln = eurpln * 1.17

    # Fetch positions with per-ticker latest price and currency from DB metadata
    pos_rows = db.execute(f"""
        SELECT p.ticker, p.shares, p.avg_cost_pln, p.account_type, p.opened_at,
               r.adj_close AS current_price_native,
               COALESCE(m.currency, 'USD') AS currency
        FROM user_positions p
        LEFT JOIN (
            SELECT rp.ticker, rp.adj_close
            FROM raw_prices rp
            INNER JOIN (
                SELECT ticker, MAX(date) AS max_date
                FROM raw_prices WHERE source = 'yfinance'
                GROUP BY ticker
            ) latest ON rp.ticker = latest.ticker AND rp.date = latest.max_date
            WHERE rp.source = 'yfinance'
        ) r ON p.ticker = r.ticker
        LEFT JOIN ticker_metadata m ON m.ticker = p.ticker
        WHERE p.user_id = '{user_id}'
        ORDER BY p.ticker
    """).fetchall()

    positions = []
    total_value = 0.0

    for r in pos_rows:
        ticker, shares, avg_cost, account_type, opened_at, price_native, ticker_currency = r

        # Convert native price to PLN using currency from ticker_metadata
        if price_native is not None:
            if ticker_currency == 'EUR':
                current_price_pln = price_native * eurpln
            elif ticker_currency in ('GBP', 'GBp'):
                current_price_pln = price_native * gbppln
            else:
                current_price_pln = price_native * usdpln
        else:
            current_price_pln = None

        value_pln = (current_price_pln * shares) if current_price_pln else None
        gain_pct  = ((current_price_pln - avg_cost) / avg_cost) if (current_price_pln and avg_cost) else None
        if value_pln:
            total_value += value_pln

        positions.append(Position(
            ticker        = ticker,
            shares        = shares,
            avg_cost_pln  = avg_cost,
            account_type  = account_type,
            opened_at     = opened_at,
            current_price = round(current_price_pln, 2) if current_price_pln else None,
            value_pln     = value_pln,
            gain_pct      = gain_pct,
        ))

    # IKE contribution status for current year
    year = date.today().year
    ike_row = db.execute(f"""
        SELECT contributed_pln, limit_pln
        FROM ike_contributions
        WHERE user_id = '{user_id}' AND year = {year}
    """).fetchone()

    contributed = ike_row[0] if ike_row else 0.0
    limit       = (ike_row[1] if ike_row and ike_row[1] else None) or IKE_LIMITS.get(year)
    remaining   = (limit - contributed) if limit else None

    # Per-account summary (IKE + IKZE + regular), aggregated live from transactions
    accounts = _compute_account_summary(db, user_id, year)

    return PortfolioResponse(
        user_id         = user_id,
        positions       = positions,
        total_value_pln = total_value or None,
        ike_contributed = contributed,
        ike_limit       = limit,
        ike_remaining   = remaining,
        accounts        = accounts,
    )


@router.post("/transaction", response_model=TransactionResponse)
def record_transaction(
    tx:      TransactionCreate,
    db:      Annotated[object, Depends(get_db_write)],
    user_id: Annotated[str, Depends(get_current_user)],
):
    require_non_guest(user_id)
    tx_id = str(uuid.uuid4())

    # Ensure user exists
    db.execute(f"""
        INSERT INTO users (user_id, email, display_name)
        VALUES ('{user_id}', 'dev@local', 'Dev User')
        ON CONFLICT (user_id) DO NOTHING
    """)

    # Validate deposit
    if tx.type == "deposit":
        if not tx.amount_pln:
            raise HTTPException(status_code=400, detail="amount_pln required for deposit")
        if tx.account_type == "regular":
            raise HTTPException(status_code=400, detail="Deposits only apply to IKE or IKZE accounts")

    # Insert transaction — use parameterized values for all user-supplied strings
    price_val = tx.amount_pln if tx.type == "deposit" else tx.price_pln
    db.execute("""
        INSERT INTO user_transactions
            (transaction_id, user_id, ticker, date, type,
             shares, price_pln, usdpln_rate, account_type, notes)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, [tx_id, user_id, tx.ticker, str(tx.date), tx.type,
          tx.shares, price_val, tx.usdpln_rate, tx.account_type, tx.notes])

    # Update positions (upsert)
    if tx.type == "deposit":
        year   = tx.date.year
        amount = tx.amount_pln
        limit  = IKE_LIMITS.get(year, 28260.0)
        db.execute(f"""
            INSERT INTO ike_contributions (user_id, year, contributed_pln, limit_pln)
            VALUES ('{user_id}', {year}, {amount}, {limit})
            ON CONFLICT (user_id, year) DO UPDATE SET
                contributed_pln = ike_contributions.contributed_pln + {amount},
                limit_pln = {limit}
        """)

    elif tx.type == "buy":
        db.execute(f"""
            INSERT INTO user_positions
                (user_id, ticker, shares, avg_cost_pln, avg_cost_usdpln,
                 account_type, opened_at)
            VALUES (
                '{user_id}', '{tx.ticker}', {tx.shares}, {tx.price_pln},
                {tx.usdpln_rate if tx.usdpln_rate else 'NULL'},
                '{tx.account_type}', '{tx.date}'
            )
            ON CONFLICT (user_id, ticker, account_type) DO UPDATE SET
                shares       = user_positions.shares + {tx.shares},
                avg_cost_pln = (user_positions.avg_cost_pln * user_positions.shares
                                + {tx.price_pln} * {tx.shares})
                               / (user_positions.shares + {tx.shares}),
                updated_at   = now()
        """)

        # Track IKE contribution
        if tx.account_type == "IKE":
            year = tx.date.year
            amount = tx.shares * tx.price_pln
            limit = IKE_LIMITS.get(year, 28260.0)
            db.execute(f"""
                INSERT INTO ike_contributions (user_id, year, contributed_pln, limit_pln)
                VALUES ('{user_id}', {year}, {amount}, {limit})
                ON CONFLICT (user_id, year) DO UPDATE SET
                    contributed_pln = ike_contributions.contributed_pln + {amount},
                    limit_pln = {limit}
            """)

    elif tx.type == "sell":
        row = db.execute(f"""
            SELECT shares FROM user_positions
            WHERE user_id = '{user_id}'
              AND ticker = '{tx.ticker}'
              AND account_type = '{tx.account_type}'
        """).fetchone()
        if not row or row[0] < tx.shares:
            raise HTTPException(status_code=400, detail="Insufficient shares to sell")
        db.execute(f"""
            UPDATE user_positions
            SET shares = shares - {tx.shares}, updated_at = now()
            WHERE user_id = '{user_id}'
              AND ticker = '{tx.ticker}'
              AND account_type = '{tx.account_type}'
        """)

    db.commit()

    if tx.type == "deposit":
        msg = f"Deposit of {tx.amount_pln:.2f} PLN to {tx.account_type} recorded."
    else:
        msg = f"{tx.type.capitalize()} of {tx.shares} {tx.ticker} recorded."

    return TransactionResponse(transaction_id=tx_id, message=msg)


@router.get("/price/{ticker}")
def get_ticker_price(
    ticker:  str,
    db:      Annotated[object, Depends(get_db)],
    _user:   Annotated[str, Depends(get_current_user)],
    on_date: str = None,  # optional YYYY-MM-DD; if omitted uses latest available
):
    """Return price for a ticker on a given date (or latest) converted to PLN.
    Uses the closest available trading day on or before the requested date."""
    # Price: closest trading day on or before requested date
    if on_date:
        price_row = db.execute(f"""
            SELECT adj_close, date FROM raw_prices
            WHERE ticker = '{ticker}' AND source = 'yfinance' AND date <= '{on_date}'
            ORDER BY date DESC LIMIT 1
        """).fetchone()
    else:
        price_row = db.execute(f"""
            SELECT adj_close, date FROM raw_prices
            WHERE ticker = '{ticker}' AND source = 'yfinance'
            ORDER BY date DESC LIMIT 1
        """).fetchone()

    if not price_row:
        raise HTTPException(status_code=404, detail=f"No price data for {ticker}" +
                            (f" on or before {on_date}" if on_date else ""))

    price_native = float(price_row[0])
    price_date   = str(price_row[1])

    # FX rate: use the rate on the same day as the price
    fx_row = db.execute(f"""
        SELECT eurpln, usdpln FROM daily_features
        WHERE eurpln IS NOT NULL AND usdpln IS NOT NULL AND date <= '{price_date}'
        ORDER BY date DESC LIMIT 1
    """).fetchone()
    eurpln = float(fx_row[0]) if fx_row else 4.25
    usdpln = float(fx_row[1]) if fx_row else 3.85
    gbppln = eurpln * 1.17

    # Look up currency from ticker_metadata (populated by ingestion)
    meta = db.execute("SELECT currency FROM ticker_metadata WHERE ticker = %s", [ticker]).fetchone()
    ticker_ccy = (meta[0] or 'USD').upper() if meta else 'USD'
    if ticker_ccy == 'EUR':
        price_pln = price_native * eurpln
        currency  = 'EUR'
    elif ticker_ccy in ('GBP', 'GBP'):
        price_pln = price_native * gbppln
        currency  = 'GBP'
    else:
        price_pln = price_native * usdpln
        currency  = 'USD'

    return {
        "ticker":        ticker,
        "price_pln":     round(price_pln, 2),
        "price_native":  round(price_native, 4),
        "currency":      currency,
        "price_date":    price_date,
    }


@router.get("/transactions", response_model=TransactionsResponse)
def get_transactions(
    db:      Annotated[object, Depends(get_db)],
    user_id: Annotated[str, Depends(get_current_user)],
    ticker:  str  = None,
    type:    str  = None,
    year:    int  = None,
):
    filters = [f"user_id = '{user_id}'"]
    if ticker:
        filters.append(f"ticker = '{ticker}'")
    if type:
        filters.append(f"type = '{type}'")
    if year:
        filters.append(f"YEAR(date) = {year}")
    where = " AND ".join(filters)

    rows = db.execute(f"""
        SELECT transaction_id, ticker, date, type, shares, price_pln,
               account_type, notes, created_at
        FROM user_transactions
        WHERE {where}
        ORDER BY date DESC, created_at DESC
    """).fetchall()

    return TransactionsResponse(transactions=[
        TransactionRow(
            transaction_id = r[0],
            ticker         = r[1],
            date           = r[2],
            type           = r[3],
            shares         = r[4],
            price_pln      = r[5],
            account_type   = r[6],
            notes          = r[7],
            created_at     = r[8],
        )
        for r in rows
    ])


@router.delete("/transaction/{transaction_id}", response_model=TransactionResponse)
def delete_transaction(
    transaction_id: str,
    db:      Annotated[object, Depends(get_db_write)],
    user_id: Annotated[str, Depends(get_current_user)],
):
    require_non_guest(user_id)
    row = db.execute(f"""
        SELECT ticker, type, shares, price_pln, account_type, date
        FROM user_transactions
        WHERE transaction_id = '{transaction_id}' AND user_id = '{user_id}'
    """).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Transaction not found")

    ticker, tx_type, shares, price_pln, account_type, tx_date = row

    db.execute(f"""
        DELETE FROM user_transactions
        WHERE transaction_id = '{transaction_id}' AND user_id = '{user_id}'
    """)

    # Reverse the position effect
    if tx_type == "buy":
        pos_row = db.execute(f"""
            SELECT shares, avg_cost_pln FROM user_positions
            WHERE user_id = '{user_id}' AND ticker = '{ticker}' AND account_type = '{account_type}'
        """).fetchone()
        if pos_row:
            new_shares = pos_row[0] - shares
            if new_shares <= 0:
                db.execute(f"""
                    DELETE FROM user_positions
                    WHERE user_id = '{user_id}' AND ticker = '{ticker}' AND account_type = '{account_type}'
                """)
            else:
                # Reverse weighted average cost
                new_avg = ((pos_row[1] * pos_row[0]) - (price_pln * shares)) / new_shares
                db.execute(f"""
                    UPDATE user_positions
                    SET shares = {new_shares}, avg_cost_pln = {new_avg}, updated_at = now()
                    WHERE user_id = '{user_id}' AND ticker = '{ticker}' AND account_type = '{account_type}'
                """)
        if account_type == "IKE":
            amount = shares * price_pln
            db.execute(f"""
                UPDATE ike_contributions
                SET contributed_pln = GREATEST(0, contributed_pln - {amount})
                WHERE user_id = '{user_id}' AND year = {tx_date.year}
            """)

    elif tx_type == "sell":
        db.execute(f"""
            UPDATE user_positions
            SET shares = shares + {shares}, updated_at = now()
            WHERE user_id = '{user_id}' AND ticker = '{ticker}' AND account_type = '{account_type}'
        """)

    elif tx_type == "deposit":
        # price_pln stores the deposit amount for deposits
        if price_pln and account_type in ("IKE", "IKZE"):
            db.execute(f"""
                UPDATE ike_contributions
                SET contributed_pln = GREATEST(0, contributed_pln - {price_pln})
                WHERE user_id = '{user_id}' AND year = {tx_date.year}
            """)

    db.commit()
    return TransactionResponse(transaction_id=transaction_id, message="Transaction deleted.")


@router.put("/transaction/{transaction_id}", response_model=TransactionResponse)
def edit_transaction(
    transaction_id: str,
    tx:      TransactionCreate,
    db:      Annotated[object, Depends(get_db_write)],
    user_id: Annotated[str, Depends(get_current_user)],
):
    require_non_guest(user_id)
    exists = db.execute(f"""
        SELECT 1 FROM user_transactions
        WHERE transaction_id = '{transaction_id}' AND user_id = '{user_id}'
    """).fetchone()
    if not exists:
        raise HTTPException(status_code=404, detail="Transaction not found")

    db.execute("""
        UPDATE user_transactions
        SET ticker = %s, date = %s, type = %s,
            shares = %s, price_pln = %s,
            account_type = %s, notes = %s
        WHERE transaction_id = %s AND user_id = %s
    """, [tx.ticker, str(tx.date), tx.type,
          tx.shares, tx.price_pln, tx.account_type, tx.notes,
          transaction_id, user_id])

    # Rebuild positions from scratch for this ticker+account_type
    _rebuild_position(db, user_id, tx.ticker, tx.account_type)

    db.commit()
    return TransactionResponse(transaction_id=transaction_id, message="Transaction updated.")


def _rebuild_position(db, user_id: str, ticker: str, account_type: str):
    """Recalculate position from all transactions for ticker+account after an edit."""
    rows = db.execute(f"""
        SELECT type, shares, price_pln, date
        FROM user_transactions
        WHERE user_id = '{user_id}' AND ticker = '{ticker}' AND account_type = '{account_type}'
        ORDER BY date ASC, created_at ASC
    """).fetchall()

    total_shares = 0.0
    total_cost   = 0.0
    opened_at    = None

    for tx_type, shares, price_pln, tx_date in rows:
        if tx_type == "buy":
            total_cost   = total_cost + shares * price_pln
            total_shares = total_shares + shares
            if opened_at is None:
                opened_at = tx_date
        elif tx_type == "sell":
            if total_shares > 0:
                avg = total_cost / total_shares
                total_cost   = total_cost - shares * avg
                total_shares = total_shares - shares

    db.execute(f"""
        DELETE FROM user_positions
        WHERE user_id = '{user_id}' AND ticker = '{ticker}' AND account_type = '{account_type}'
    """)

    if total_shares > 0:
        avg_cost = total_cost / total_shares
        db.execute(f"""
            INSERT INTO user_positions (user_id, ticker, shares, avg_cost_pln, account_type, opened_at)
            VALUES ('{user_id}', '{ticker}', {total_shares}, {avg_cost}, '{account_type}', '{opened_at}')
        """)


@router.get("/template")
def download_template(_user: Annotated[str, Depends(get_current_user)]):
    """Return a pre-filled .xlsx template for bulk transaction import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["date", "ticker", "type", "shares", "price_pln", "account_type", "notes"]
    ws.append(headers)

    # Example row so the user understands the format
    ws.append(["2024-01-15", "VWCE.DE", "buy", 1.5, 410.50, "IKE", "Monthly DCA"])

    # Column widths
    for col, width in zip("ABCDEFG", [14, 12, 8, 10, 12, 14, 30]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=transactions_template.xlsx"},
    )


@router.post("/upload")
def upload_transactions(
    file:    UploadFile = File(...),
    db:      Annotated[object, Depends(get_db_write)] = None,
    user_id: Annotated[str, Depends(get_current_user)] = None,
):
    """Bulk-import transactions from an .xlsx file (same template as /portfolio/template)."""
    require_non_guest(user_id)
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    try:
        contents = file.file.read()
        df = pd.read_excel(io.BytesIO(contents), dtype=str)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read file: {e}")

    required = {"date", "ticker", "type", "shares", "price_pln"}
    missing = required - set(df.columns.str.strip().str.lower())
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(missing)}")

    df.columns = df.columns.str.strip().str.lower()

    imported, errors = 0, []

    # Ensure user exists
    db.execute(f"""
        INSERT INTO users (user_id, email, display_name)
        VALUES ('{user_id}', 'dev@local', 'Dev User')
        ON CONFLICT (user_id) DO NOTHING
    """)

    for i, row in df.iterrows():
        row_num = i + 2  # 1-indexed, header is row 1

        # --- validate ---
        try:
            tx_date = pd.to_datetime(row["date"]).date()
        except Exception:
            errors.append(f"Row {row_num}: invalid date '{row.get('date')}'")
            continue

        if tx_date > date.today():
            errors.append(f"Row {row_num}: date {tx_date} is in the future")
            continue

        ticker = str(row["ticker"]).strip().upper()
        tx_type = str(row["type"]).strip().lower()
        if tx_type not in ("buy", "sell", "dividend"):
            errors.append(f"Row {row_num}: type must be buy/sell/dividend, got '{tx_type}'")
            continue

        try:
            shares = float(row["shares"])
            price_pln = float(row["price_pln"])
        except Exception:
            errors.append(f"Row {row_num}: shares and price_pln must be numbers")
            continue

        if shares <= 0 or price_pln <= 0:
            errors.append(f"Row {row_num}: shares and price_pln must be > 0")
            continue

        account_type = str(row.get("account_type", "IKE")).strip()
        if account_type not in ("IKE", "IKZE", "regular"):
            account_type = "IKE"

        notes = str(row.get("notes", "")).strip()
        notes_sql = f"'{notes}'" if notes and notes != "nan" else "NULL"

        tx_id = str(uuid.uuid4())

        # --- insert transaction ---
        db.execute(f"""
            INSERT INTO user_transactions
                (transaction_id, user_id, ticker, date, type,
                 shares, price_pln, account_type, notes)
            VALUES (
                '{tx_id}', '{user_id}', '{ticker}', '{tx_date}', '{tx_type}',
                {shares}, {price_pln}, '{account_type}', {notes_sql}
            )
        """)

        # --- update positions ---
        if tx_type == "buy":
            db.execute(f"""
                INSERT INTO user_positions
                    (user_id, ticker, shares, avg_cost_pln, account_type, opened_at)
                VALUES ('{user_id}', '{ticker}', {shares}, {price_pln}, '{account_type}', '{tx_date}')
                ON CONFLICT (user_id, ticker, account_type) DO UPDATE SET
                    shares       = user_positions.shares + {shares},
                    avg_cost_pln = (user_positions.avg_cost_pln * user_positions.shares
                                    + {price_pln} * {shares})
                                   / (user_positions.shares + {shares}),
                    updated_at   = now()
            """)
            if account_type == "IKE":
                amount = shares * price_pln
                limit = IKE_LIMITS.get(tx_date.year, 26019.0)
                db.execute(f"""
                    INSERT INTO ike_contributions (user_id, year, contributed_pln, limit_pln)
                    VALUES ('{user_id}', {tx_date.year}, {amount}, {limit})
                    ON CONFLICT (user_id, year) DO UPDATE SET
                        contributed_pln = ike_contributions.contributed_pln + {amount},
                        limit_pln = {limit}
                """)

        elif tx_type == "sell":
            pos = db.execute(f"""
                SELECT shares FROM user_positions
                WHERE user_id = '{user_id}' AND ticker = '{ticker}' AND account_type = '{account_type}'
            """).fetchone()
            if not pos or pos[0] < shares:
                errors.append(f"Row {row_num}: insufficient shares to sell {shares} {ticker}")
                continue
            db.execute(f"""
                UPDATE user_positions
                SET shares = shares - {shares}, updated_at = now()
                WHERE user_id = '{user_id}' AND ticker = '{ticker}' AND account_type = '{account_type}'
            """)

        imported += 1

    db.commit()

    return {
        "imported": imported,
        "errors":   errors,
        "message":  f"{imported} transaction(s) imported." + (f" {len(errors)} row(s) skipped." if errors else ""),
    }


_BROKER_GEMINI_MODEL = "gemini-3.1-flash-lite"
_MAX_UPLOAD_BYTES    = 5 * 1024 * 1024  # 5 MB hard limit
# xlsx magic bytes: PK\x03\x04 (ZIP format)
_XLSX_MAGIC = b"PK\x03\x04"
# csv: no magic — accepted by extension only, content validated downstream
_ALLOWED_EXTENSIONS = {".xlsx", ".csv"}

_BROKER_PROMPT = """You are a data extraction assistant. The user has uploaded a broker transaction export.
Your task: map the columns to this schema and return ONLY a JSON array of transaction objects.

There are TWO types of valid rows:

Type 1 — stock transaction:
{
  "date": "YYYY-MM-DD",
  "ticker": "TICKER.EXCHANGE (Yahoo Finance format, e.g. VWCE.DE)",
  "type": "buy" | "sell" | "dividend",
  "shares": <positive number>,
  "price_native": <positive number — price per share in the instrument's native currency>,
  "currency": "EUR" | "USD" | "GBP" | "PLN",
  "account_type": "regular",
  "notes": "<optional string or null>"
}

Type 2 — cash deposit to investment account (IKE, IKZE, or similar):
{
  "date": "YYYY-MM-DD",
  "type": "deposit",
  "amount_pln": <positive number — deposit amount in PLN>,
  "account_type": "regular",
  "notes": "<optional string or null>"
}

For deposit rows: look for rows with type/description like "IKE Deposit", "IKZE Deposit", "Transfer in", "Deposit", "Wpłata" with a positive PLN amount and NO ticker/instrument.

DO NOT convert prices to PLN — return the native price and currency. The backend will apply the correct historical FX rate.

Ticker normalisation rules:
- Any ticker ending in .UK → replace .UK with .L (e.g. ISAC.UK → ISAC.L, EGLN.UK → EGLN.L)
- Tickers ending in .DE, .L, .PA stay as-is
- Strip exchange prefix if present (e.g. "LSE:ISAC" → "ISAC.L")
- If currency is PLN, ticker is likely already on Warsaw Stock Exchange — keep as-is

Other rules:
- Always set account_type to "regular" — the user will change it in the UI
- Ignore rows that are fees, taxes, currency exchanges, or stock purchase/sale cash movements (those are captured by the buy/sell rows)
- If you cannot determine a required field, omit that row
- Return ONLY the JSON array. No explanation, no markdown fences, no other text."""


def _sanitise_cell(val) -> str:
    """Strip non-printable characters from a cell value to prevent prompt injection."""
    s = str(val) if val is not None else ""
    return re.sub(r"[^\x20-\x7E -￿]", "", s)[:200]


@router.post("/upload-broker")
async def upload_broker(
    file:       UploadFile = File(...),
    db:         Annotated[object, Depends(get_db_write)] = None,
    user_id:    Annotated[str, Depends(get_current_user)] = None,
    sheet_name: str = None,  # if None: return sheet list; if set: parse that sheet
):
    """AI-assisted broker export import. Accepts any broker .xlsx or .csv format.
    First call (no sheet_name): returns available sheets.
    Second call (sheet_name set): parses that sheet with Gemini and returns preview."""
    require_non_guest(user_id)
    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(status_code=503, detail="GEMINI_API_KEY not configured.")

    # --- Security: extension check ---
    suffix = "." + (file.filename or "").rsplit(".", 1)[-1].lower()
    if suffix not in _ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Only .xlsx or .csv files are accepted.")

    # --- Security: size limit ---
    contents = await file.read()
    if len(contents) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=400, detail="File too large (max 5 MB).")

    # --- Security: magic bytes check for xlsx ---
    if suffix == ".xlsx" and not contents[:4] == _XLSX_MAGIC:
        raise HTTPException(status_code=400, detail="File does not appear to be a valid .xlsx.")

    # --- For xlsx: return sheet names if no sheet selected yet ---
    if suffix == ".xlsx":
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not parse file: {e}")

        if not sheet_name:
            # First call — just return sheet names, no AI call yet
            return {"sheets": sheet_names, "preview": None, "parse_errors": [], "message": "Select a sheet to parse."}

        # Second call — load the requested sheet
        if sheet_name not in sheet_names:
            raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found. Available: {sheet_names}")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            ws = wb[sheet_name]
            rows_raw = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read sheet: {e}")
    else:
        # CSV — single sheet, skip sheet selection
        try:
            import csv
            text = contents.decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows_raw = list(reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not parse file: {e}")

    if len(rows_raw) < 2:
        raise HTTPException(status_code=400, detail="Sheet appears empty.")

    pre_extracted_deposits = []  # reserved for future deposit parsing

    # --- Build sanitised preview for Gemini (headers + first 30 data rows) ---
    header_row = [_sanitise_cell(c) for c in (rows_raw[0] or [])]
    data_rows  = [
        [_sanitise_cell(c) for c in (row or [])]
        for row in rows_raw[1:31]  # max 30 rows sent to model
    ]

    # Fetch latest FX rates to help the model convert prices
    fx_row = db.execute("""
        SELECT eurpln, usdpln FROM daily_features
        WHERE eurpln IS NOT NULL AND usdpln IS NOT NULL
        ORDER BY date DESC LIMIT 1
    """).fetchone()
    eur_pln = float(fx_row[0]) if fx_row else 4.25
    usd_pln = float(fx_row[1]) if fx_row else 3.85

    table_text = "Headers: " + ", ".join(header_row) + "\n"
    for row in data_rows:
        table_text += " | ".join(row) + "\n"

    prompt = (
        f"Current FX rates: EUR/PLN = {eur_pln:.4f}, USD/PLN = {usd_pln:.4f}\n\n"
        + table_text + "\n" + _BROKER_PROMPT
    )

    # --- Call Gemini ---
    try:
        from google import genai
        client_ai = genai.Client(api_key=gemini_key)
        response = client_ai.models.generate_content(
            model=_BROKER_GEMINI_MODEL,
            contents=prompt,
        )
        raw_json = response.text.strip()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"AI parsing failed: {e}")

    # --- Parse and strictly validate Gemini output ---
    # Strip markdown fences if model added them despite instructions
    raw_json = re.sub(r"^```(?:json)?\s*", "", raw_json)
    raw_json = re.sub(r"\s*```$", "", raw_json)

    try:
        parsed = json.loads(raw_json)
        if not isinstance(parsed, list):
            raise ValueError("Expected a JSON array")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"AI returned invalid JSON: {e}. Raw: {raw_json[:300]}")

    # Pre-fetch all EUR/PLN and USD/PLN rates we might need (one query, keyed by date string)
    fx_rows = db.execute("""
        SELECT date::text, eurpln, usdpln, gbppln
        FROM (
            SELECT date,
                   eurpln,
                   usdpln,
                   eurpln * 1.17 AS gbppln
            FROM daily_features
            WHERE eurpln IS NOT NULL AND usdpln IS NOT NULL
        ) t
    """).fetchall()
    # Build dict: date_str -> (eurpln, usdpln, gbppln)
    fx_by_date = {r[0]: (float(r[1]), float(r[2]), float(r[3])) for r in fx_rows}

    def _get_fx(tx_date_str: str, currency: str) -> float:
        """Return PLN per 1 unit of currency, using closest available date."""
        # Find closest date on or before tx_date
        rate = fx_by_date.get(tx_date_str)
        if not rate:
            # Walk back up to 10 days to find closest trading day
            from datetime import timedelta
            d = date.fromisoformat(tx_date_str)
            for i in range(1, 11):
                key = (d - timedelta(days=i)).isoformat()
                if key in fx_by_date:
                    rate = fx_by_date[key]
                    break
        if not rate:
            rate = (4.25, 3.85, 4.97)  # hard fallback
        eur, usd, gbp = rate
        if currency == "EUR":   return eur
        if currency == "USD":   return usd
        if currency in ("GBP", "GBp"): return gbp
        return 1.0  # PLN — no conversion needed

    today = date.today()
    validated, parse_errors = [], []

    for i, item in enumerate(parsed):
        row_num = i + 1
        if not isinstance(item, dict):
            parse_errors.append(f"Row {row_num}: not an object"); continue

        # date
        try:
            tx_date = date.fromisoformat(str(item.get("date", ""))[:10])
        except Exception:
            parse_errors.append(f"Row {row_num}: invalid date '{item.get('date')}'"); continue
        if tx_date > today:
            parse_errors.append(f"Row {row_num}: date {tx_date} is in the future"); continue

        # type
        tx_type = str(item.get("type", "")).strip().lower()
        if tx_type not in ("buy", "sell", "dividend", "deposit"):
            parse_errors.append(f"Row {row_num}: invalid type '{tx_type}'"); continue

        # Handle deposit rows separately — no ticker/shares/price needed
        if tx_type == "deposit":
            try:
                amount_pln = float(str(item.get("amount_pln", 0)).replace(",", "."))
            except Exception:
                parse_errors.append(f"Row {row_num}: amount_pln must be a number"); continue
            if amount_pln <= 0:
                parse_errors.append(f"Row {row_num}: amount_pln must be > 0"); continue
            account_type = str(item.get("account_type", "IKE")).strip()
            if account_type not in ("IKE", "IKZE", "regular"):
                account_type = "IKE"
            notes_raw = item.get("notes") or ""
            notes = re.sub(r"[^\x20-\x7E]", "", str(notes_raw))[:200] or None
            validated.append({
                "date": tx_date.isoformat(), "type": "deposit",
                "amount_pln": amount_pln, "account_type": account_type, "notes": notes,
            })
            continue

        # ticker — alphanumeric + dot + dash only
        ticker = str(item.get("ticker", "")).strip().upper()
        if not re.fullmatch(r"[A-Z0-9.\-]{1,20}", ticker):
            parse_errors.append(f"Row {row_num}: invalid ticker '{ticker}'"); continue

        # shares
        try:
            shares = float(str(item.get("shares", 0)).replace(",", "."))
        except Exception:
            parse_errors.append(f"Row {row_num}: shares must be a number"); continue
        if shares <= 0:
            parse_errors.append(f"Row {row_num}: shares must be > 0"); continue

        # price_native + currency → convert to PLN using historical FX
        currency = str(item.get("currency", "EUR")).strip().upper()
        if currency not in ("EUR", "USD", "GBP", "GBp", "PLN"):
            currency = "EUR"  # safe default for European ETFs

        # Override AI-guessed currency with the actual currency from ticker_metadata.
        # Fixes cases like ISAC.L which trades on LSE (looks like GBP) but is USD-denominated.
        meta_row = db.execute(
            "SELECT currency FROM ticker_metadata WHERE ticker = %s", [ticker]
        ).fetchone()
        if meta_row and meta_row[0]:
            currency = meta_row[0].upper()

        try:
            price_native = float(str(item.get("price_native", 0)).replace(",", "."))
        except Exception:
            parse_errors.append(f"Row {row_num}: price_native must be a number"); continue
        if price_native <= 0:
            parse_errors.append(f"Row {row_num}: price_native must be > 0"); continue

        fx_rate = _get_fx(tx_date.isoformat(), currency)
        price_pln = round(price_native * fx_rate, 4)

        # account_type
        account_type = str(item.get("account_type", "regular")).strip()
        if account_type not in ("IKE", "IKZE", "regular"):
            account_type = "regular"

        # notes — strip to plain text
        notes_raw = item.get("notes") or ""
        notes = re.sub(r"[^\x20-\x7E]", "", str(notes_raw))[:200] or None

        validated.append({
            "date": tx_date.isoformat(), "ticker": ticker, "type": tx_type,
            "shares": shares, "price_pln": price_pln,
            "price_native": price_native, "currency": currency,
            "account_type": account_type, "notes": notes,
        })

    # Return preview — caller must POST /portfolio/upload-broker/confirm to commit
    return {
        "preview":      validated,
        "parse_errors": parse_errors,
        "message":      f"AI parsed {len(validated)} transaction(s). Review and confirm.",
        "total_rows_sent_to_ai": len(data_rows),
    }


@router.post("/upload-broker/confirm")
def confirm_broker_import(
    payload: dict,
    db:      Annotated[object, Depends(get_db_write)] = None,
    user_id: Annotated[str, Depends(get_current_user)] = None,
):
    """Commit a previously previewed broker import. Expects {transactions: [...]}."""
    require_non_guest(user_id)
    rows = payload.get("transactions", [])
    if not rows:
        raise HTTPException(status_code=400, detail="No transactions provided.")
    if len(rows) > 500:
        raise HTTPException(status_code=400, detail="Too many transactions (max 500 per import).")

    today = date.today()
    imported, errors = [], []

    db.execute("""
        INSERT INTO users (user_id, email, display_name)
        VALUES (%s, 'dev@local', 'Dev User')
        ON CONFLICT (user_id) DO NOTHING
    """, [user_id])

    for i, item in enumerate(rows):
        row_num = i + 1
        # Re-validate every field — never trust client-supplied data
        try:
            tx_date = date.fromisoformat(str(item.get("date", "")))
        except Exception:
            errors.append(f"Row {row_num}: invalid date"); continue
        if tx_date > today:
            errors.append(f"Row {row_num}: date in the future"); continue

        tx_type = str(item.get("type", "")).strip().lower()
        if tx_type not in ("buy", "sell", "dividend", "deposit"):
            errors.append(f"Row {row_num}: invalid type"); continue

        # Deposit rows handled separately
        if tx_type == "deposit":
            try:
                amount_pln = float(item.get("amount_pln", 0))
            except Exception:
                errors.append(f"Row {row_num}: invalid amount_pln"); continue
            if amount_pln <= 0:
                errors.append(f"Row {row_num}: amount_pln must be > 0"); continue
            account_type = str(item.get("account_type", "IKE")).strip()
            if account_type not in ("IKE", "IKZE", "regular"):
                account_type = "IKE"
            notes_raw = item.get("notes") or ""
            notes = re.sub(r"[^\x20-\x7E]", "", str(notes_raw))[:200] or None
            tx_id = str(uuid.uuid4())
            db.execute("""
                INSERT INTO user_transactions
                    (transaction_id, user_id, ticker, date, type,
                     shares, price_pln, account_type, notes)
                VALUES (%s, %s, NULL, %s, 'deposit', NULL, %s, %s, %s)
            """, [tx_id, user_id, tx_date, amount_pln, account_type, notes])
            if account_type in ("IKE", "IKZE"):
                limit = IKE_LIMITS.get(tx_date.year, 28260.0)
                db.execute("""
                    INSERT INTO ike_contributions (user_id, year, contributed_pln, limit_pln)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (user_id, year) DO UPDATE SET
                        contributed_pln = ike_contributions.contributed_pln + EXCLUDED.contributed_pln,
                        limit_pln = EXCLUDED.limit_pln
                """, [user_id, tx_date.year, amount_pln, limit])
            imported.append(tx_id)
            continue

        ticker = str(item.get("ticker", "")).strip().upper()
        if not re.fullmatch(r"[A-Z0-9.\-]{1,20}", ticker):
            errors.append(f"Row {row_num}: invalid ticker"); continue

        try:
            shares    = float(item.get("shares", 0))
            price_pln = float(item.get("price_pln", 0))
        except Exception:
            errors.append(f"Row {row_num}: invalid shares/price"); continue
        if shares <= 0 or price_pln <= 0:
            errors.append(f"Row {row_num}: shares/price must be > 0"); continue

        account_type = str(item.get("account_type", "IKE")).strip()
        if account_type not in ("IKE", "IKZE", "regular"):
            account_type = "IKE"

        notes_raw = item.get("notes") or ""
        notes = re.sub(r"[^\x20-\x7E]", "", str(notes_raw))[:200] or None

        tx_id = str(uuid.uuid4())
        notes_param = notes if notes else None

        db.execute("""
            INSERT INTO user_transactions
                (transaction_id, user_id, ticker, date, type,
                 shares, price_pln, account_type, notes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, [tx_id, user_id, ticker, tx_date, tx_type,
              shares, price_pln, account_type, notes_param])

        if tx_type == "buy":
            db.execute("""
                INSERT INTO user_positions
                    (user_id, ticker, shares, avg_cost_pln, account_type, opened_at)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (user_id, ticker, account_type) DO UPDATE SET
                    shares       = user_positions.shares + EXCLUDED.shares,
                    avg_cost_pln = (user_positions.avg_cost_pln * user_positions.shares
                                    + EXCLUDED.avg_cost_pln * EXCLUDED.shares)
                                   / (user_positions.shares + EXCLUDED.shares),
                    updated_at   = now()
            """, [user_id, ticker, shares, price_pln, account_type, tx_date])

            if account_type == "IKE":
                amount = shares * price_pln
                limit  = IKE_LIMITS.get(tx_date.year, 26019.0)
                db.execute("""
                    INSERT INTO ike_contributions (user_id, year, contributed_pln, limit_pln)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (user_id, year) DO UPDATE SET
                        contributed_pln = ike_contributions.contributed_pln + EXCLUDED.contributed_pln,
                        limit_pln = EXCLUDED.limit_pln
                """, [user_id, tx_date.year, amount, limit])

        elif tx_type == "sell":
            pos = db.execute("""
                SELECT shares FROM user_positions
                WHERE user_id = %s AND ticker = %s AND account_type = %s
            """, [user_id, ticker, account_type]).fetchone()
            if not pos or pos[0] < shares:
                errors.append(f"Row {row_num}: insufficient shares to sell {shares} {ticker}")
                continue
            db.execute("""
                UPDATE user_positions
                SET shares = shares - %s, updated_at = now()
                WHERE user_id = %s AND ticker = %s AND account_type = %s
            """, [shares, user_id, ticker, account_type])

        imported.append(tx_id)

    db.commit()
    return {
        "imported": len(imported),
        "errors":   errors,
        "message":  f"{len(imported)} transaction(s) imported." + (f" {len(errors)} skipped." if errors else ""),
    }


@router.delete("/transactions/all", response_model=TransactionResponse)
def delete_all_transactions(
    db:      Annotated[object, Depends(get_db_write)],
    user_id: Annotated[str, Depends(get_current_user)],
):
    """Delete all transactions and positions for the current user."""
    require_non_guest(user_id)
    db.execute("DELETE FROM user_transactions WHERE user_id = %s", [user_id])
    db.execute("DELETE FROM user_positions WHERE user_id = %s", [user_id])
    db.execute("DELETE FROM ike_contributions WHERE user_id = %s", [user_id])
    db.commit()
    return TransactionResponse(transaction_id="", message="All transactions and positions deleted.")


@router.get("/analysis")
def get_portfolio_analysis(
    db:      Annotated[object, Depends(get_db)],
    user_id: Annotated[str, Depends(get_current_user)],
):
    """Return weighted region + sector + commodity breakdown across all held ETF positions.

    Logic:
      1. Fetch current positions with PLN values (same FX logic as get_portfolio).
      2. For each position, look up etf_allocations for that ticker.
      3. Weight each region/sector/commodity allocation by the position's PLN value.
      4. Normalise totals to 100%.
      5. Return {regions, sectors, commodities, total_value_pln, coverage_pct}.

    coverage_pct = fraction of total portfolio value that has allocation data.
    Gold/commodity ETFs contribute only to the 'commodities' breakdown (not regions/sectors).
    """
    # FX rates
    fx_row = db.execute("""
        SELECT eurpln, usdpln FROM daily_features
        WHERE eurpln IS NOT NULL AND usdpln IS NOT NULL
        ORDER BY date DESC LIMIT 1
    """).fetchone()
    eurpln = float(fx_row[0]) if fx_row else 4.25
    usdpln = float(fx_row[1]) if fx_row else 3.85
    gbppln = eurpln * 1.17

    # Positions with current price and currency
    pos_rows = db.execute(f"""
        SELECT p.ticker, p.shares,
               r.adj_close AS current_price_native,
               COALESCE(m.currency, 'USD') AS currency
        FROM user_positions p
        LEFT JOIN (
            SELECT rp.ticker, rp.adj_close
            FROM raw_prices rp
            INNER JOIN (
                SELECT ticker, MAX(date) AS max_date
                FROM raw_prices WHERE source = 'yfinance'
                GROUP BY ticker
            ) latest ON rp.ticker = latest.ticker AND rp.date = latest.max_date
            WHERE rp.source = 'yfinance'
        ) r ON p.ticker = r.ticker
        LEFT JOIN ticker_metadata m ON m.ticker = p.ticker
        WHERE p.user_id = '{user_id}'
    """).fetchall()

    if not pos_rows:
        return {
            "regions": [], "sectors": [], "commodities": [],
            "total_value_pln": 0.0, "coverage_pct": 0.0,
        }

    # Compute PLN value per position
    position_values = {}  # ticker -> value_pln (may aggregate multiple account_types)
    for ticker, shares, price_native, ticker_currency in pos_rows:
        if price_native is None:
            continue
        if ticker_currency == 'EUR':
            price_pln = float(price_native) * eurpln
        elif ticker_currency in ('GBP', 'GBp'):
            price_pln = float(price_native) * gbppln
        else:
            price_pln = float(price_native) * usdpln
        val = float(shares) * price_pln
        position_values[ticker] = position_values.get(ticker, 0.0) + val

    total_value = sum(position_values.values())
    if total_value <= 0:
        return {
            "regions": [], "sectors": [], "commodities": [],
            "total_value_pln": 0.0, "coverage_pct": 0.0,
        }

    # Fetch allocation data for all held tickers in one query
    tickers_held = list(position_values.keys())
    placeholders = ", ".join(["%s"] * len(tickers_held))
    alloc_rows = db.execute(
        f"SELECT ticker, allocation_type, label, weight FROM etf_allocations WHERE ticker IN ({placeholders})",
        tickers_held,
    ).fetchall()

    # Build lookup: ticker -> {alloc_type -> {label -> weight}}
    alloc_map: dict = {}
    for ticker, alloc_type, label, weight in alloc_rows:
        alloc_map.setdefault(ticker, {}).setdefault(alloc_type, {})[label] = float(weight)

    # Weighted aggregation
    region_totals:    dict[str, float] = {}
    sector_totals:    dict[str, float] = {}
    commodity_totals: dict[str, float] = {}
    covered_value = 0.0

    for ticker, val in position_values.items():
        ticker_allocs = alloc_map.get(ticker)
        if not ticker_allocs:
            continue  # no data for this ticker — skip (reduces coverage_pct)
        covered_value += val
        weight_fraction = val / total_value  # how much of portfolio this ticker is

        for label, w in ticker_allocs.get("region", {}).items():
            region_totals[label] = region_totals.get(label, 0.0) + w * weight_fraction

        for label, w in ticker_allocs.get("sector", {}).items():
            sector_totals[label] = sector_totals.get(label, 0.0) + w * weight_fraction

        for label, w in ticker_allocs.get("commodity", {}).items():
            commodity_totals[label] = commodity_totals.get(label, 0.0) + w * weight_fraction

    # Normalise to covered portion only, then express as % of portfolio
    # (weights already sum correctly if ETF weights sum to 1.0)
    def _to_sorted_list(totals: dict) -> list:
        if not totals:
            return []
        items = sorted(totals.items(), key=lambda x: x[1], reverse=True)
        return [{"label": lbl, "weight_pct": round(w * 100, 2)} for lbl, w in items]

    coverage_pct = round(covered_value / total_value * 100, 1) if total_value > 0 else 0.0

    return {
        "regions":         _to_sorted_list(region_totals),
        "sectors":         _to_sorted_list(sector_totals),
        "commodities":     _to_sorted_list(commodity_totals),
        "total_value_pln": round(total_value, 2),
        "coverage_pct":    coverage_pct,
    }


@router.post("/allocations/refresh")
def refresh_allocations(
    db:      Annotated[object, Depends(get_db_write)],
    _user:   Annotated[str, Depends(get_current_user)],
):
    """Refresh ETF allocation data from iShares public API.

    Rate-limited: only runs if the most recent allocation row is older than 7 days.
    If iShares is unreachable the seed data remains unchanged.
    Returns {"updated": [...], "failed": [...], "skipped": [...], "skipped_rate_limit": bool}.
    """
    require_non_guest(_user)
    import datetime as _dt
    import logging
    _log = logging.getLogger(__name__)

    # Rate-limit check: only refresh if last update was >7 days ago
    row = db.execute("""
        SELECT MAX(updated_at) FROM etf_allocations
    """).fetchone()
    last_updated = row[0] if row else None

    if last_updated is not None:
        now_utc = _dt.datetime.now(_dt.timezone.utc)
        if hasattr(last_updated, "tzinfo") and last_updated.tzinfo is not None:
            age = now_utc - last_updated
        else:
            age = now_utc - last_updated.replace(tzinfo=_dt.timezone.utc)
        if age.days < 7:
            _log.info("Allocation refresh skipped — last update was %s days ago", age.days)
            return {
                "updated": [],
                "failed":  [],
                "skipped": [],
                "skipped_rate_limit": True,
                "message": (
                    f"Allocation data is up to date (last updated {age.days} day(s) ago). "
                    "Refresh allowed every 7 days."
                ),
            }

    from backend.etf_ishares_fetch import refresh_allocations_from_ishares
    result = refresh_allocations_from_ishares(db)
    result["skipped_rate_limit"] = False

    if result["updated"]:
        result["message"] = (
            f"Updated {len(result['updated'])} ticker(s): {', '.join(result['updated'])}."
            + (f" Failed: {', '.join(result['failed'])}." if result["failed"] else "")
        )
    else:
        result["message"] = (
            "No tickers updated from iShares. Seed data unchanged."
            + (f" Failed: {', '.join(result['failed'])}." if result["failed"] else "")
        )

    return result


@router.get("/ike-history")
def get_ike_history(
    db:      Annotated[object, Depends(get_db)],
    user_id: Annotated[str, Depends(get_current_user)],
):
    """Return IKE contribution history for all years, with limits filled in."""
    rows = db.execute(f"""
        SELECT year, contributed_pln, limit_pln
        FROM ike_contributions
        WHERE user_id = '{user_id}'
        ORDER BY year ASC
    """).fetchall()

    # Merge DB rows with known limits; fill gaps for years with no contributions
    result = []
    years_in_db = {r[0] for r in rows}
    db_by_year  = {r[0]: r for r in rows}

    # Include all years from 2016 up to current year
    current_year = date.today().year
    for year in range(2016, current_year + 1):
        limit = IKE_LIMITS.get(year)
        if year in years_in_db:
            _, contributed, db_limit = db_by_year[year]
            result.append({
                "year":        year,
                "contributed": round(contributed or 0.0, 2),
                "limit":       round(db_limit or limit or 0.0, 2),
                "pct":         round((contributed or 0) / (db_limit or limit or 1) * 100, 1),
            })
        elif limit:
            result.append({
                "year": year, "contributed": 0.0,
                "limit": limit, "pct": 0.0,
            })

    return {"years": result}
